from django.core.mail import send_mail
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.models import Role
from accounts.permissions import OwnerOrStaffWritePermission, RolePermission

from .csv_utils import CsvMixin, parse_bool, parse_int
from .filters import AlumniFilter, ReferralLeadFilter
from .pdf import render_student_brochure
from .models import (
    Alumni,
    AlumniSubmission,
    AuditLog,
    Company,
    Event,
    EventParticipant,
    JobIntelResponse,
    JobPosting,
    MessageTemplate,
    Notification,
    OutreachCampaign,
    OutreachContact,
    ReferralLead,
    Student,
    Task,
)
from .notifications import notify_all_users, notify_users
from .serializers import (
    AlumniSerializer,
    AlumniSubmissionSerializer,
    AuditLogSerializer,
    CompanySerializer,
    EventParticipantSerializer,
    EventSerializer,
    JobIntelResponseSerializer,
    JobPostingSerializer,
    MessageTemplateSerializer,
    NotificationSerializer,
    OutreachCampaignSerializer,
    OutreachContactSerializer,
    ReferralLeadSerializer,
    StudentSerializer,
    TaskSerializer,
)


class AuditedModelViewSet(viewsets.ModelViewSet):
    """Base ViewSet: RBAC on every endpoint + writes go to the audit log."""

    permission_classes = [RolePermission]
    audit_entity = None  # override
    # Char fields whose values the UI may add/delete on the fly (free-text lists).
    editable_choice_fields = []

    @action(detail=False, methods=["post"], url_path="delete-value")
    def delete_value(self, request):
        """Remove a custom value from a free-text choice field by reassigning
        every record that uses it to `reassign_to` (a safe default value).
        Admin-only."""
        if not getattr(request.user, "is_admin", False):
            return Response({"detail": "Only admins can delete values."}, status=403)
        field = request.data.get("field")
        if field not in self.editable_choice_fields:
            return Response({"detail": "field not editable"}, status=400)
        value = (request.data.get("value") or "").strip()
        reassign_to = request.data.get("reassign_to") or ""
        if not value:
            return Response({"detail": "value is required"}, status=400)
        model = self.queryset.model
        n = model.objects.filter(**{field: value}).update(**{field: reassign_to})
        return Response({"reassigned": n})

    def _log(self, action_name, instance):
        AuditLog.objects.create(
            user=self.request.user if self.request.user.is_authenticated else None,
            action=action_name,
            entity=self.audit_entity or self.queryset.model.__name__,
            entity_id=str(getattr(instance, "pk", "")),
            summary=str(instance)[:255],
        )

    def perform_create(self, serializer):
        instance = serializer.save()
        self._log("created", instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        self._log("updated", instance)

    def perform_destroy(self, instance):
        self._log("deleted", instance)
        instance.delete()


class CompanyViewSet(AuditedModelViewSet):
    queryset = Company.objects.all()
    serializer_class = CompanySerializer
    audit_entity = "Company"
    search_fields = ["name", "sector"]
    filterset_fields = ["is_in_placement_list", "sector"]
    ordering_fields = ["name", "created_at"]

    def get_queryset(self):
        # The directory listing is driven by alumni employers: only surface
        # companies that have at least one linked alumnus (PRD §7 companies).
        # Other actions (retrieve/update/destroy) operate on all companies so a
        # company can still be deleted even after its last alumnus leaves it.
        qs = Company.objects.annotate(num_alumni=Count("alumni"))
        if self.action == "list":
            return qs.filter(num_alumni__gt=0)
        return qs

    @action(detail=False, methods=["get"])
    def names(self, request):
        """All companies (unfiltered) as {id, name} for autocomplete suggestions
        and delete-from-dropdown when assigning a company to an alumnus."""
        return Response(
            list(Company.objects.order_by("name").values("id", "name"))
        )


class AlumniViewSet(CsvMixin, AuditedModelViewSet):
    queryset = Alumni.objects.select_related("company", "updated_by").all()
    serializer_class = AlumniSerializer
    audit_entity = "Alumni"
    editable_choice_fields = ["status", "role_level"]
    filterset_class = AlumniFilter
    search_fields = ["name", "email", "domain", "city", "company__name"]
    ordering_fields = ["name", "batch", "willingness", "updated_at"]

    csv_filename = "alumni"
    csv_columns = [
        "name", "batch", "branch", "company", "role_level", "domain",
        "city", "email", "phone", "linkedin", "source", "referred_by",
        "status", "is_super_alumni", "willingness",
    ]

    def perform_create(self, serializer):
        instance = serializer.save(updated_by=self.request.user)
        self._log("created", instance)

    def perform_update(self, serializer):
        instance = serializer.save(updated_by=self.request.user)
        self._log("updated", instance)

    @action(detail=False, methods=["get"])
    def branches(self, request):
        """Distinct branches currently in use, so the UI can offer them as
        suggestions alongside newly added ones."""
        values = (
            Alumni.objects.exclude(branch="")
            .values_list("branch", flat=True)
            .distinct()
        )
        return Response(sorted(set(values)))

    @action(detail=False, methods=["get"])
    def batches(self, request):
        """Distinct graduation years, newest first, for the filter dropdown."""
        values = Alumni.objects.values_list("batch", flat=True).distinct()
        return Response(sorted({v for v in values if v}, reverse=True))

    @action(detail=False, methods=["post"])
    def delete_branch(self, request):
        """Remove a (custom) branch by reassigning everyone on it to 'Other'.
        Applies to both alumni and students since they share the branch space.
        Admin-only."""
        if not getattr(request.user, "is_admin", False):
            return Response({"detail": "Only admins can delete branches."}, status=403)
        branch = (request.data.get("branch") or "").strip()
        if not branch:
            return Response({"detail": "branch is required"}, status=400)
        if branch == "Other":
            return Response({"detail": "'Other' cannot be deleted"}, status=400)
        reassigned_alumni = Alumni.objects.filter(branch=branch).update(branch="Other")
        reassigned_students = Student.objects.filter(branch=branch).update(branch="Other")
        return Response(
            {"reassigned_alumni": reassigned_alumni, "reassigned_students": reassigned_students}
        )

    def row_to_dict(self, obj):
        return {
            "name": obj.name, "batch": obj.batch, "branch": obj.branch,
            "company": obj.company.name if obj.company else "",
            "role_level": obj.role_level, "domain": obj.domain, "city": obj.city,
            "email": obj.email, "phone": obj.phone, "linkedin": obj.linkedin,
            "source": obj.source, "referred_by": obj.referred_by,
            "status": obj.status, "is_super_alumni": obj.is_super_alumni,
            "willingness": obj.willingness,
        }

    def import_row(self, row):
        email = (row.get("email") or "").strip()
        if not email:
            raise ValueError("email is required")
        # Only update columns actually present in the CSV — a partial file must
        # not blank out fields it doesn't include.
        defaults = {"updated_by": self.request.user}
        if "name" in row:
            defaults["name"] = (row["name"] or "").strip()
        if "batch" in row:
            defaults["batch"] = parse_int(row["batch"], 0)
        if "branch" in row:
            defaults["branch"] = (row["branch"] or "Other").strip()
        if "company" in row:
            cname = (row["company"] or "").strip()
            defaults["company"] = (
                Company.objects.get_or_create(name=cname)[0] if cname else None
            )
        if "role_level" in row:
            defaults["role_level"] = (row["role_level"] or "").strip()
        if "domain" in row:
            defaults["domain"] = (row["domain"] or "").strip()
        if "city" in row:
            defaults["city"] = (row["city"] or "").strip()
        if "phone" in row:
            defaults["phone"] = (row["phone"] or "").strip()
        if "linkedin" in row:
            defaults["linkedin"] = (row["linkedin"] or "").strip()
        if "source" in row:
            defaults["source"] = (row["source"] or "").strip()
        if "referred_by" in row:
            defaults["referred_by"] = (row["referred_by"] or "").strip()
        if "status" in row:
            defaults["status"] = (row["status"] or "active").strip()
        if "is_super_alumni" in row:
            defaults["is_super_alumni"] = parse_bool(row["is_super_alumni"])
        if "willingness" in row:
            defaults["willingness"] = parse_int(row["willingness"], 3)
        return Alumni.objects.update_or_create(email=email, defaults=defaults)

    @action(detail=False, methods=["get"])
    def segments(self, request):
        """Quick segmentation summary — counts by branch and by city."""
        by_branch = list(
            Alumni.objects.values("branch").annotate(count=Count("id")).order_by("-count")
        )
        by_city = list(
            Alumni.objects.exclude(city="")
            .values("city")
            .annotate(count=Count("id"))
            .order_by("-count")[:10]
        )
        return Response({"by_branch": by_branch, "by_city": by_city})


class AlumniSubmissionViewSet(
    mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet
):
    """Review queue for self-service alumni submissions (public form).
    Staff list pending items and approve (apply to directory) / reject them."""

    queryset = AlumniSubmission.objects.select_related("reviewed_by").all()
    serializer_class = AlumniSubmissionSerializer
    permission_classes = [RolePermission]
    write_roles = {Role.ADMIN, Role.COORDINATOR, Role.VOLUNTEER}
    filterset_fields = ["status"]
    ordering_fields = ["created_at", "status"]

    def _apply_to_directory(self, sub, user):
        cname = (sub.company or "").strip()
        company = Company.objects.get_or_create(name=cname)[0] if cname else None
        existing = Alumni.objects.filter(email__iexact=sub.email).first()
        if existing:
            existing.name = sub.name
            existing.batch = sub.batch
            existing.branch = sub.branch
            existing.consent_given = True
            if cname:
                existing.company = company
            for key in ("role_level", "domain", "city", "phone", "linkedin", "source", "referred_by", "photo"):
                if getattr(sub, key):
                    setattr(existing, key, getattr(sub, key))
            existing.updated_by = user
            existing.save()
            return existing, False
        alum = Alumni.objects.create(
            name=sub.name, email=sub.email, batch=sub.batch, branch=sub.branch,
            company=company, role_level=sub.role_level, domain=sub.domain,
            city=sub.city, phone=sub.phone, linkedin=sub.linkedin,
            source=sub.source, referred_by=sub.referred_by, photo=sub.photo,
            status=Alumni.Status.ACTIVE, consent_given=True, updated_by=user,
        )
        return alum, True

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        sub = self.get_object()
        if sub.status != AlumniSubmission.Status.PENDING:
            return Response({"detail": "This submission was already reviewed."}, status=400)
        alum, created = self._apply_to_directory(sub, request.user)
        sub.status = AlumniSubmission.Status.APPROVED
        sub.reviewed_by = request.user
        sub.save(update_fields=["status", "reviewed_by", "updated_at"])
        AuditLog.objects.create(
            user=request.user, action="approved", entity="AlumniSubmission",
            entity_id=str(sub.pk),
            summary=f"{'Added' if created else 'Updated'} {sub.name} <{sub.email}>",
        )
        return Response({"detail": f"Approved — alumnus {'added' if created else 'updated'}.", "alumni_id": alum.id})

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        sub = self.get_object()
        if sub.status != AlumniSubmission.Status.PENDING:
            return Response({"detail": "This submission was already reviewed."}, status=400)
        sub.status = AlumniSubmission.Status.REJECTED
        sub.reviewed_by = request.user
        sub.save(update_fields=["status", "reviewed_by", "updated_at"])
        AuditLog.objects.create(
            user=request.user, action="rejected", entity="AlumniSubmission",
            entity_id=str(sub.pk), summary=f"Rejected {sub.name} <{sub.email}>",
        )
        return Response({"detail": "Submission rejected."})


class StudentViewSet(CsvMixin, AuditedModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    audit_entity = "Student"
    editable_choice_fields = ["branch"]
    search_fields = ["name", "domain", "email"]
    filterset_fields = ["branch", "batch"]
    ordering_fields = ["name", "batch", "gpa"]

    csv_filename = "students"
    csv_columns = ["name", "batch", "branch", "gpa", "domain", "skills", "email"]

    def row_to_dict(self, obj):
        skills = "; ".join(obj.skills) if isinstance(obj.skills, list) else ""
        return {
            "name": obj.name, "batch": obj.batch, "branch": obj.branch,
            "gpa": obj.gpa or "", "domain": obj.domain, "skills": skills,
            "email": obj.email,
        }

    def import_row(self, row):
        name = (row.get("name") or "").strip()
        if not name:
            raise ValueError("name is required")
        batch = parse_int(row.get("batch"), 0)
        # Only update columns present in the CSV (partial files don't blank fields).
        defaults = {}
        if "branch" in row:
            defaults["branch"] = (row["branch"] or "Other").strip()
        if "gpa" in row:
            defaults["gpa"] = row["gpa"] or None
        if "domain" in row:
            defaults["domain"] = (row["domain"] or "").strip()
        if "skills" in row:
            raw_skills = (row["skills"] or "").replace(",", ";")
            defaults["skills"] = [s.strip() for s in raw_skills.split(";") if s.strip()]
        if "email" in row:
            defaults["email"] = (row["email"] or "").strip()
        return Student.objects.update_or_create(
            name=name, batch=batch, defaults=defaults
        )

    @action(detail=True, methods=["get"])
    def brochure(self, request, pk=None):
        """Auto-generate the two-page talent brochure as a PDF (PRD §3)."""
        return render_student_brochure(self.get_object())


class MessageTemplateViewSet(AuditedModelViewSet):
    queryset = MessageTemplate.objects.all()
    serializer_class = MessageTemplateSerializer
    audit_entity = "MessageTemplate"
    search_fields = ["name", "type"]
    filterset_fields = ["channel"]


class OutreachCampaignViewSet(AuditedModelViewSet):
    queryset = OutreachCampaign.objects.select_related("owner", "template").all()
    serializer_class = OutreachCampaignSerializer
    audit_entity = "OutreachCampaign"
    editable_choice_fields = ["channel"]
    search_fields = ["name"]
    filterset_fields = ["channel", "owner"]

    def perform_create(self, serializer):
        instance = serializer.save(owner=self.request.user)
        self._log("created", instance)

    def get_throttles(self):
        # Rate-limit the outreach-send endpoint per PRD §10.
        if self.action == "send":
            self.throttle_scope = "outreach_send"
        return super().get_throttles()

    @action(detail=True, methods=["post"])
    def populate(self, request, pk=None):
        """Add alumni matching the campaign's saved segment as pending contacts."""
        campaign = self.get_object()
        seg = campaign.segment_filter or {}
        qs = Alumni.objects.filter(status=Alumni.Status.ACTIVE)
        if seg.get("branch"):
            qs = qs.filter(branch=seg["branch"])
        if seg.get("city"):
            qs = qs.filter(city__icontains=seg["city"])
        if seg.get("domain"):
            qs = qs.filter(domain__icontains=seg["domain"])
        if seg.get("min_willingness"):
            qs = qs.filter(willingness__gte=int(seg["min_willingness"]))
        added = 0
        for alum in qs:
            _, created = OutreachContact.objects.get_or_create(
                campaign=campaign, alumni=alum,
                defaults={"status": OutreachContact.Status.PENDING},
            )
            added += int(created)
        self._log("updated", campaign)
        return Response({"matched": qs.count(), "added": added})

    @action(detail=True, methods=["post"])
    def send(self, request, pk=None):
        """Send the campaign to all pending contacts.

        Email channel sends real email (via the configured backend — console in
        dev, SendGrid/SES SMTP in prod). WhatsApp/LinkedIn touches are only
        logged, not auto-sent (PRD §8). Rate limited per PRD §10.
        """
        campaign = self.get_object()
        body_tmpl = campaign.template.body if campaign.template else "Hello {name},"
        contacts = campaign.contacts.filter(
            status=OutreachContact.Status.PENDING
        ).select_related("alumni")
        sent = 0
        for contact in contacts:
            alum = contact.alumni
            if campaign.channel == "email" and alum.email:
                message = (
                    body_tmpl.replace("{name}", alum.name)
                    .replace("{company}", alum.company.name if alum.company else "")
                    .replace("{domain}", alum.domain or "")
                )
                send_mail(
                    subject=campaign.name,
                    message=message,
                    from_email=None,
                    recipient_list=[alum.email],
                    fail_silently=True,
                )
            contact.status = OutreachContact.Status.SENT
            contact.sent_at = timezone.now()
            contact.save(update_fields=["status", "sent_at", "updated_at"])
            sent += 1
        self._log("updated", campaign)
        return Response({"sent": sent, "channel": campaign.channel})


class OutreachContactViewSet(AuditedModelViewSet):
    queryset = OutreachContact.objects.select_related("alumni", "campaign").all()
    serializer_class = OutreachContactSerializer
    audit_entity = "OutreachContact"
    filterset_fields = ["campaign", "status", "alumni"]


class EventViewSet(AuditedModelViewSet):
    queryset = Event.objects.all()
    serializer_class = EventSerializer
    audit_entity = "Event"
    editable_choice_fields = ["type"]
    search_fields = ["title", "venue"]
    filterset_fields = ["type"]
    ordering_fields = ["date", "title"]

    def _announce_event(self, instance, verb="New event"):
        if instance.date and instance.date >= timezone.now():
            from .tasks import send_event_announcement
            try:
                send_event_announcement.delay(instance.id)
            except Exception:
                send_event_announcement(instance.id)
            when = timezone.localtime(instance.date).strftime("%d %b %Y, %I:%M %p")
            notify_all_users(
                title=f"{verb}: {instance.title}",
                message=f"{when} · {instance.venue or 'TBA'}",
                link="/events",
                kind=Notification.Kind.EVENT,
            )

    def perform_create(self, serializer):
        instance = serializer.save()
        self._log("created", instance)
        self._announce_event(instance, verb="New event")

    def perform_update(self, serializer):
        old_date = serializer.instance.date
        instance = serializer.save()
        self._log("updated", instance)
        # Notify if date changed to a future time
        if instance.date != old_date:
            self._announce_event(instance, verb="Event updated")


class EventParticipantViewSet(AuditedModelViewSet):
    queryset = EventParticipant.objects.select_related("event").all()
    serializer_class = EventParticipantSerializer
    audit_entity = "EventParticipant"
    filterset_fields = ["event", "person_type", "role", "rsvp", "attended"]
    # Volunteers may RSVP people in / mark attendance.
    write_roles = {Role.ADMIN, Role.COORDINATOR, Role.VOLUNTEER}


class ReferralLeadViewSet(AuditedModelViewSet):
    queryset = ReferralLead.objects.select_related("alumni", "student", "company").all()
    serializer_class = ReferralLeadSerializer
    audit_entity = "ReferralLead"
    editable_choice_fields = ["stage", "outcome"]
    filterset_class = ReferralLeadFilter
    ordering_fields = ["created_at", "last_followup_at", "stage"]
    write_roles = {Role.ADMIN, Role.COORDINATOR, Role.VOLUNTEER}

    @action(detail=True, methods=["post"])
    def followup(self, request, pk=None):
        """Record a follow-up — resets the 48h SLA clock."""
        lead = self.get_object()
        lead.last_followup_at = timezone.now()
        lead.save(update_fields=["last_followup_at", "updated_at"])
        self._log("updated", lead)
        return Response(ReferralLeadSerializer(lead).data)


class JobIntelResponseViewSet(AuditedModelViewSet):
    queryset = JobIntelResponse.objects.select_related("alumni").all()
    serializer_class = JobIntelResponseSerializer
    audit_entity = "JobIntelResponse"
    editable_choice_fields = ["timeline"]
    filterset_fields = ["hiring", "timeline", "alumni"]
    # Built-in form: alumni can submit their own pulse.
    write_roles = {Role.ADMIN, Role.COORDINATOR, Role.VOLUNTEER, Role.ALUMNUS}


class TaskViewSet(AuditedModelViewSet):
    queryset = Task.objects.select_related("assignee").all()
    serializer_class = TaskSerializer
    audit_entity = "Task"
    editable_choice_fields = ["team", "status"]
    filterset_fields = ["team", "status", "assignee"]
    ordering_fields = ["due_date", "created_at"]
    write_roles = {Role.ADMIN, Role.COORDINATOR, Role.VOLUNTEER}

    def _notify_assignee(self, task, title):
        # Ping the assignee — but never notify someone about their own action.
        if task.assignee_id and task.assignee_id != self.request.user.id:
            notify_users(
                [task.assignee],
                title=title,
                message=task.title,
                link="/tasks",
                kind=Notification.Kind.TASK,
            )

    def perform_create(self, serializer):
        instance = serializer.save()
        self._log("created", instance)
        self._notify_assignee(instance, "You have a new task")

    def perform_update(self, serializer):
        prev_assignee = serializer.instance.assignee_id
        instance = serializer.save()
        self._log("updated", instance)
        if instance.assignee_id != prev_assignee:
            self._notify_assignee(instance, "A task was assigned to you")


class JobPostingViewSet(AuditedModelViewSet):
    queryset = JobPosting.objects.select_related("posted_by").all()
    serializer_class = JobPostingSerializer
    audit_entity = "JobPosting"
    editable_choice_fields = ["work_mode", "employment_type"]
    permission_classes = [OwnerOrStaffWritePermission]
    search_fields = ["title", "company", "location"]
    filterset_fields = ["work_mode", "employment_type", "is_open"]
    ordering_fields = ["created_at", "title"]
    # Alumni can post jobs too; any logged-in member can read them.
    write_roles = {Role.ADMIN, Role.COORDINATOR, Role.VOLUNTEER, Role.ALUMNUS}

    def perform_create(self, serializer):
        instance = serializer.save(posted_by=self.request.user)
        self._log("created", instance)
        # The job board is visible to everyone → broadcast new openings.
        notify_all_users(
            title=f"New job: {instance.title}",
            message=f"{instance.company} · {instance.location or 'Remote'}",
            link="/jobs",
            kind=Notification.Kind.JOB,
            exclude=self.request.user,
        )

    @action(detail=False, methods=["get"])
    def stats(self, request):
        """Aggregate figures for the Jobs & Internships header + sidebar."""
        qs = JobPosting.objects.all()
        month_start = timezone.now().replace(
            day=1, hour=0, minute=0, second=0, microsecond=0
        )
        total = qs.count()
        internships = qs.filter(
            employment_type=JobPosting.EmploymentType.INTERNSHIP
        ).count()
        companies = qs.values("company").distinct().count()
        open_jobs = qs.filter(is_open=True).count()
        new_this_month = qs.filter(created_at__gte=month_start).count()

        top_companies = list(
            qs.values("company")
            .annotate(openings=Count("id"))
            .order_by("-openings")[:5]
        )
        return Response(
            {
                "total": total,
                "internships": internships,
                "companies": companies,
                "open_jobs": open_jobs,
                "new_this_month": new_this_month,
                "top_companies": top_companies,
            }
        )


class NotificationViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """Each user's own notification inbox — list, mark read, dismiss."""

    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ["is_read", "kind"]
    ordering_fields = ["created_at"]

    def get_queryset(self):
        return Notification.objects.filter(recipient=self.request.user)

    @action(detail=False, methods=["get"])
    def unread_count(self, request):
        return Response({"unread": self.get_queryset().filter(is_read=False).count()})

    @action(detail=False, methods=["post"])
    def mark_all_read(self, request):
        updated = self.get_queryset().filter(is_read=False).update(is_read=True)
        return Response({"marked": updated})

    @action(detail=True, methods=["post"])
    def read(self, request, pk=None):
        notif = self.get_object()
        if not notif.is_read:
            notif.is_read = True
            notif.save(update_fields=["is_read", "updated_at"])
        return Response(self.get_serializer(notif).data)


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only audit trail — admins and coordinators only."""

    queryset = AuditLog.objects.select_related("user").all()
    serializer_class = AuditLogSerializer
    permission_classes = [RolePermission]
    read_roles = {Role.ADMIN, Role.COORDINATOR}
    filterset_fields = ["action", "entity", "user"]
    search_fields = ["entity", "summary"]


class DashboardView(APIView):
    """KPI analytics (PRD §3 Dashboards) — aggregate counts across modules."""

    permission_classes = [RolePermission]

    def get(self, request):
        from datetime import timedelta

        from django.contrib.auth import get_user_model
        from django.db.models.functions import TruncMonth

        from .models import JobPosting

        sla_cutoff = timezone.now() - timedelta(hours=48)
        breached = ReferralLead.objects.filter(
            outcome=ReferralLead.Outcome.PENDING
        ).exclude(stage=ReferralLead.Stage.CLOSED).filter(
            Q(last_followup_at__lt=sla_cutoff)
            | Q(last_followup_at__isnull=True, created_at__lt=sla_cutoff)
        ).count()

        my_tasks_open = Task.objects.filter(
            assignee=request.user
        ).exclude(status="done").count() if request.user.is_authenticated else 0

        # Alumni growth: cumulative month-by-month for current year
        now = timezone.now()
        year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        monthly_new = (
            Alumni.objects
            .filter(created_at__gte=year_start)
            .annotate(month=TruncMonth("created_at"))
            .values("month")
            .annotate(count=Count("id"))
            .order_by("month")
        )
        base_count = Alumni.objects.filter(created_at__lt=year_start).count()
        months_short = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        growth_map = {r["month"].month: r["count"] for r in monthly_new}
        alumni_growth = []
        running = base_count
        for m in range(1, now.month + 1):
            running += growth_map.get(m, 0)
            alumni_growth.append({"month": months_short[m - 1], "alumni": running})

        # Recent activities from audit log
        recent_logs = AuditLog.objects.select_related("user").order_by("-timestamp")[:8]
        recent_activities = [
            {
                "id": log.pk,
                "action": log.action,
                "entity": log.entity,
                "summary": log.summary,
                "user": log.user.name if log.user and hasattr(log.user, "name") else (log.user.email if log.user else "System"),
                "timestamp": log.timestamp.isoformat(),
            }
            for log in recent_logs
        ]

        # ── month-over-month growth (real %) ──
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if month_start.month == 1:
            prev_month_start = month_start.replace(year=month_start.year - 1, month=12)
        else:
            prev_month_start = month_start.replace(month=month_start.month - 1)

        def mom_pct(model):
            this_m = model.objects.filter(created_at__gte=month_start).count()
            prev_m = model.objects.filter(
                created_at__gte=prev_month_start, created_at__lt=month_start
            ).count()
            if prev_m:
                return round(((this_m - prev_m) / prev_m) * 100, 1)
            return None

        # ── alumni by location (top cities) ──
        loc_rows = list(
            Alumni.objects.exclude(city="").values("city")
            .annotate(count=Count("id")).order_by("-count")
        )
        by_location = [{"label": r["city"], "value": r["count"]} for r in loc_rows[:6]]
        other_loc = sum(r["count"] for r in loc_rows[6:])
        if other_loc:
            by_location.append({"label": "Others", "value": other_loc})

        # ── events this month (with RSVP counts) ──
        month_events = (
            Event.objects.filter(date__gte=month_start)
            .annotate(pcount=Count("participants"))
            .order_by("date")[:5]
        )
        events_this_month = [
            {
                "id": e.id, "title": e.title, "type": e.type,
                "date": e.date.isoformat() if e.date else None,
                "venue": e.venue, "participant_count": e.pcount,
            }
            for e in month_events
        ]

        # ── top hiring companies ──
        top_companies = list(
            JobPosting.objects.values("company")
            .annotate(openings=Count("id")).order_by("-openings")[:5]
        )

        # ── jobs by employment type (donut) ──
        type_labels = dict(JobPosting.EmploymentType.choices)
        jobs_by_type = [
            {"label": type_labels.get(r["employment_type"], r["employment_type"]), "value": r["count"]}
            for r in JobPosting.objects.values("employment_type")
            .annotate(count=Count("id")).order_by("-count")
        ]

        data = {
            "jobs_open": JobPosting.objects.filter(is_open=True).count(),
            "users_total": get_user_model().objects.count(),
            "my_tasks_open": my_tasks_open,
            "alumni_total": Alumni.objects.count(),
            "alumni_active": Alumni.objects.filter(status="active").count(),
            "super_alumni": Alumni.objects.filter(is_super_alumni=True).count(),
            "jobs_total": JobPosting.objects.count(),
            "alumni_delta_pct": mom_pct(Alumni),
            "events_delta_pct": mom_pct(Event),
            "jobs_delta_pct": mom_pct(JobPosting),
            "alumni_by_location": by_location,
            "events_this_month": events_this_month,
            "top_companies": top_companies,
            "jobs_by_type": jobs_by_type,
            "students_total": Student.objects.count(),
            "companies_total": Company.objects.count(),
            "companies_in_placement": Company.objects.filter(is_in_placement_list=True).count(),
            "campaigns_total": OutreachCampaign.objects.count(),
            "outreach_sent": OutreachContact.objects.filter(status="sent").count(),
            "outreach_replied": OutreachContact.objects.filter(status="replied").count(),
            "events_total": Event.objects.count(),
            "events_upcoming": Event.objects.filter(date__gte=timezone.now()).count(),
            "referrals_open": ReferralLead.objects.exclude(stage="closed").count(),
            "referrals_placed": ReferralLead.objects.filter(outcome="placed").count(),
            "referrals_sla_breached": breached,
            "tasks_open": Task.objects.exclude(status="done").count(),
            "hiring_now": JobIntelResponse.objects.filter(hiring=True).count(),
            "referrals_by_stage": list(
                ReferralLead.objects.values("stage").annotate(count=Count("id")).order_by("stage")
            ),
            "alumni_by_branch": list(
                Alumni.objects.values("branch").annotate(count=Count("id")).order_by("-count")
            ),
            "alumni_growth": alumni_growth,
            "recent_activities": recent_activities,
        }
        return Response(data)


MONTHS_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


def resolve_report_range(request):
    """Turn ?period / ?start / ?end query params into a [start, end) window.

    `period`: this_month | this_year | last_year | all  (default this_year).
    `start`/`end`: explicit YYYY-MM-DD dates (inclusive), override `period`.
    Returns (start, end, label) — start may be None only internally; callers
    get a concrete start floored to the earliest data row for "all".
    """
    from datetime import datetime, timedelta

    now = timezone.now()
    tz = now.tzinfo

    def parse_day(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").replace(tzinfo=tz)
        except (ValueError, TypeError):
            return None

    start = parse_day(request.query_params.get("start"))
    end_raw = parse_day(request.query_params.get("end"))
    end = (end_raw + timedelta(days=1)) if end_raw else None

    if start or end:
        label = "Custom range"
    else:
        period = (request.query_params.get("period") or "this_year").lower()
        if period == "this_month":
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            label = now.strftime("%B %Y")
        elif period == "last_year":
            start = now.replace(year=now.year - 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            end = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            label = str(now.year - 1)
        elif period == "all":
            start = None
            label = "All time"
        else:  # this_year
            start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            label = str(now.year)

    if end is None:
        end = now + timedelta(days=1)
    if start is None:
        first = Alumni.objects.order_by("created_at").values_list("created_at", flat=True).first()
        start = first or end.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)

    return start, end, label


def build_analytics(start, end):
    """Compute the full Reports & Analytics payload for the window [start, end)."""
    from datetime import timedelta

    from django.db.models.functions import TruncMonth

    end_incl = end - timedelta(seconds=1)
    multi_year = start.year != end_incl.year

    # list of (year, month) buckets spanning the window
    buckets, y, m = [], start.year, start.month
    while (y, m) <= (end_incl.year, end_incl.month):
        buckets.append((y, m))
        m += 1
        if m > 12:
            m, y = 1, y + 1

    def blabel(y, m):
        return MONTHS_SHORT[m - 1] + (f" '{y % 100:02d}" if multi_year else "")

    def monthly_map(qs):
        rows = (
            qs.filter(created_at__gte=start, created_at__lt=end)
            .annotate(month=TruncMonth("created_at"))
            .values("month")
            .annotate(count=Count("id"))
        )
        return {(r["month"].year, r["month"].month): r["count"] for r in rows}

    def pct(new_this, prior):
        return round((new_this / prior) * 100, 1) if prior else None

    # population as of the end of the window
    alumni_qs = Alumni.objects.filter(created_at__lt=end)

    # new-in-window vs equal-length prior window (real growth %)
    prior_start = start - (end - start)

    def new_vs_prior(model):
        cur = model.objects.filter(created_at__gte=start, created_at__lt=end).count()
        prev = model.objects.filter(created_at__gte=prior_start, created_at__lt=start).count()
        return pct(cur, prev)

    kpis = {
        "alumni_total": alumni_qs.count(),
        "alumni_growth_pct": new_vs_prior(Alumni),
        "alumni_active": alumni_qs.filter(status="active").count(),
        "events_total": Event.objects.filter(created_at__lt=end).count(),
        "events_growth_pct": new_vs_prior(Event),
        "jobs_total": JobPosting.objects.filter(created_at__lt=end).count(),
        "jobs_growth_pct": new_vs_prior(JobPosting),
        "companies_total": Company.objects.filter(created_at__lt=end).count(),
        "super_alumni": alumni_qs.filter(is_super_alumni=True).count(),
    }

    # cumulative alumni growth across the window
    growth_map = monthly_map(Alumni.objects)
    base = Alumni.objects.filter(created_at__lt=start).count()
    alumni_growth, running = [], base
    for (yy, mm) in buckets:
        running += growth_map.get((yy, mm), 0)
        alumni_growth.append({"month": blabel(yy, mm), "alumni": running})

    # by location (top cities)
    loc_rows = list(
        alumni_qs.exclude(city="").values("city")
        .annotate(count=Count("id")).order_by("-count")
    )
    by_location = [{"label": r["city"], "value": r["count"]} for r in loc_rows[:6]]
    other_loc = sum(r["count"] for r in loc_rows[6:])
    if other_loc:
        by_location.append({"label": "Others", "value": other_loc})

    # by industry (domain)
    norm = {}
    for r in alumni_qs.values("domain").annotate(count=Count("id")):
        key = (r["domain"] or "Other").strip() or "Other"
        norm[key] = norm.get(key, 0) + r["count"]
    ind_sorted = sorted(norm.items(), key=lambda kv: -kv[1])
    by_industry = [{"label": k, "value": v} for k, v in ind_sorted[:5]]
    other_ind = sum(v for _, v in ind_sorted[5:])
    if other_ind:
        by_industry.append({"label": "Others", "value": other_ind})

    # by branch
    by_branch = [
        {"label": r["branch"], "value": r["count"]}
        for r in alumni_qs.values("branch").annotate(count=Count("id")).order_by("-count")
    ]

    # engagement over time (events + jobs per month)
    ev, jb = monthly_map(Event.objects), monthly_map(JobPosting.objects)
    engagement = [
        {"month": blabel(yy, mm),
         "events": ev.get((yy, mm), 0),
         "jobs": jb.get((yy, mm), 0)}
        for (yy, mm) in buckets
    ]

    # top active alumni (by willingness to help)
    top_alumni = [
        {
            "id": a.id, "name": a.name,
            "role_level": a.role_level,
            "company": a.company.name if a.company else "",
            "willingness": a.willingness,
            "is_super_alumni": a.is_super_alumni,
        }
        for a in alumni_qs.select_related("company").order_by(
            "-is_super_alumni", "-willingness", "name"
        )[:5]
    ]

    return {
        "kpis": kpis,
        "alumni_growth": alumni_growth,
        "by_location": by_location,
        "by_industry": by_industry,
        "by_branch": by_branch,
        "engagement": engagement,
        "top_alumni": top_alumni,
    }


class AnalyticsView(APIView):
    """Reports & Analytics — richer aggregates for the reporting dashboard.

    Accepts ?period=this_month|this_year|last_year|all or ?start=&end= dates.
    """

    permission_classes = [RolePermission]
    read_roles = {Role.ADMIN, Role.COORDINATOR, Role.VOLUNTEER}

    def get(self, request):
        start, end, label = resolve_report_range(request)
        data = build_analytics(start, end)
        data["range"] = {
            "label": label,
            "start": start.date().isoformat(),
            "end": (end - timezone.timedelta(days=1)).date().isoformat(),
        }
        return Response(data)


class AnalyticsExportView(APIView):
    """Excel (.xlsx) export of the Reports & Analytics data for a date range.

    Same query params as AnalyticsView (period / start / end)."""

    permission_classes = [RolePermission]
    read_roles = {Role.ADMIN, Role.COORDINATOR, Role.VOLUNTEER}

    def get(self, request):
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        start, end, label = resolve_report_range(request)
        d = build_analytics(start, end)
        k = d["kpis"]
        start_s = start.date().isoformat()
        end_s = (end - timezone.timedelta(days=1)).date().isoformat()

        # ── shared styles ──
        TITLE_FONT = Font(bold=True, size=15, color="1E293B")
        SUB_FONT = Font(italic=True, size=9, color="64748B")
        HEAD_FONT = Font(bold=True, color="FFFFFF")
        HEAD_FILL = PatternFill("solid", fgColor="4F46E5")
        TOTAL_FONT = Font(bold=True, color="1E293B")
        TOTAL_FILL = PatternFill("solid", fgColor="EEF2FF")
        thin = Side(style="thin", color="E2E8F0")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
        CENTER = Alignment(horizontal="center", vertical="center")

        wb = Workbook()
        pct = lambda v: "N/A" if v is None else f"{v}%"

        def styled_sheet(title, headers, rows, first=False, totals=None):
            """Write a titled, colour-headed, auto-width table with borders."""
            ws = wb.active if first else wb.create_sheet()
            ws.title = title[:31]
            ncol = len(headers)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
            tc = ws.cell(row=1, column=1, value=title)
            tc.font = TITLE_FONT
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
            sc = ws.cell(row=2, column=1, value=f"{label}  ·  {start_s} → {end_s}")
            sc.font = SUB_FONT
            ws.row_dimensions[1].height = 22

            hrow = 4
            for j, h in enumerate(headers, 1):
                c = ws.cell(row=hrow, column=j, value=h)
                c.font = HEAD_FONT
                c.fill = HEAD_FILL
                c.alignment = CENTER
                c.border = BORDER
            for i, row in enumerate(rows, hrow + 1):
                for j, val in enumerate(row, 1):
                    c = ws.cell(row=i, column=j, value=val)
                    c.border = BORDER
                    if j > 1 and isinstance(val, (int, float)):
                        c.alignment = CENTER
            if totals:
                trow = hrow + 1 + len(rows)
                for j, val in enumerate(totals, 1):
                    c = ws.cell(row=trow, column=j, value=val)
                    c.font = TOTAL_FONT
                    c.fill = TOTAL_FILL
                    c.border = BORDER
                    if j > 1 and isinstance(val, (int, float)):
                        c.alignment = CENTER

            # auto width from content
            all_rows = rows + ([totals] if totals else [])
            for j, h in enumerate(headers, 1):
                width = len(str(h))
                for row in all_rows:
                    if j - 1 < len(row):
                        width = max(width, len(str(row[j - 1])))
                ws.column_dimensions[get_column_letter(j)].width = min(46, max(12, width + 3))
            ws.freeze_panes = ws.cell(row=hrow + 1, column=1)
            return ws

        # ── extra headline numbers (make the summary feel complete) ──
        students_total = Student.objects.filter(created_at__lt=end).count()
        companies_placement = Company.objects.filter(
            created_at__lt=end, is_in_placement_list=True
        ).count()
        referrals_open = ReferralLead.objects.filter(created_at__lt=end).exclude(stage="closed").count()
        referrals_placed = ReferralLead.objects.filter(created_at__lt=end, outcome="placed").count()
        campaigns_total = OutreachCampaign.objects.filter(created_at__lt=end).count()
        events_upcoming = Event.objects.filter(date__gte=timezone.now()).count()

        styled_sheet("Summary", ["Metric", "Value"], [
            ["Report period", label],
            ["From", start_s],
            ["To", end_s],
            ["Total Alumni", k["alumni_total"]],
            ["Alumni growth (vs previous period)", pct(k["alumni_growth_pct"])],
            ["Active Alumni", k["alumni_active"]],
            ["Super Alumni", k["super_alumni"]],
            ["Total Students", students_total],
            ["Events Organized", k["events_total"]],
            ["Upcoming Events", events_upcoming],
            ["Events growth (vs previous period)", pct(k["events_growth_pct"])],
            ["Jobs Posted", k["jobs_total"]],
            ["Jobs growth (vs previous period)", pct(k["jobs_growth_pct"])],
            ["Companies", k["companies_total"]],
            ["Companies in placement list", companies_placement],
            ["Open Referrals", referrals_open],
            ["Placed Referrals", referrals_placed],
            ["Outreach Campaigns", campaigns_total],
        ], first=True)

        # cumulative growth + new-per-month derived from the cumulative series
        growth_rows, prev = [], None
        for r in d["alumni_growth"]:
            new = r["alumni"] - prev if prev is not None else r["alumni"]
            growth_rows.append([r["month"], r["alumni"], max(0, new)])
            prev = r["alumni"]
        styled_sheet("Alumni Growth", ["Month", "Cumulative Alumni", "New this month"], growth_rows)

        def share_sheet(title, head, items):
            total = sum(i["value"] for i in items) or 1
            rows = [[i["label"], i["value"], f"{round(i['value'] / total * 100)}%"] for i in items]
            styled_sheet(title, [head, "Alumni", "Share"], rows,
                         totals=["Total", sum(i["value"] for i in items), "100%"])

        share_sheet("By Location", "Location", d["by_location"])
        share_sheet("By Industry", "Industry", d["by_industry"])
        share_sheet("By Branch", "Branch", d["by_branch"])

        styled_sheet("Engagement", ["Month", "Events", "Jobs"],
                     [[r["month"], r["events"], r["jobs"]] for r in d["engagement"]],
                     totals=["Total",
                             sum(r["events"] for r in d["engagement"]),
                             sum(r["jobs"] for r in d["engagement"])])

        styled_sheet("Top Alumni", ["Name", "Role", "Company", "Willingness (1-5)", "Super Alumni"],
                     [[a["name"], a["role_level"] or "—", a["company"] or "—", a["willingness"],
                       "Yes" if a["is_super_alumni"] else "No"] for a in d["top_alumni"]])

        # ── full alumni roster within the period (the real "data") ──
        roster = (
            Alumni.objects.filter(created_at__lt=end)
            .select_related("company").order_by("name")
        )
        roster_rows = [
            [
                a.name, a.batch, a.branch,
                a.company.name if a.company else "—",
                a.role_level or "—", a.domain or "—", a.city or "—",
                a.email, a.phone or "—",
                a.source or "—", a.referred_by or "—",
                a.status,
            ]
            for a in roster
        ]
        styled_sheet(
            "All Alumni",
            ["Name", "Batch", "Branch", "Company", "Role", "Domain", "City",
             "Email", "Phone", "Source", "Referred by", "Status"],
            roster_rows,
        )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"analytics-report_{start_s}_to_{end_s}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp
